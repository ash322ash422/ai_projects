"""
Builds the output .xlsx from a validated TenderData record using
openpyxl, with light formatting so the POC output looks presentable.
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models.tender_model import TenderData


class ExcelService:
    HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    LABEL_FONT = Font(bold=True)
    WRAP = Alignment(wrap_text=True, vertical="top")

    def build(self, tender: TenderData, source_filename: str = "") -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Tender Summary"

        ws["A1"] = "Field"
        ws["B1"] = "Value"
        for cell in ("A1", "B1"):
            ws[cell].fill = self.HEADER_FILL
            ws[cell].font = self.HEADER_FONT

        data = tender.model_dump()
        row = 2
        for field_key, label in TenderData.FIELD_LABELS.items():
            ws.cell(row=row, column=1, value=label).font = self.LABEL_FONT
            value_cell = ws.cell(row=row, column=2, value=data.get(field_key) or "")
            value_cell.alignment = self.WRAP
            row += 1

        if source_filename:
            ws.cell(row=row + 1, column=1, value="Source File").font = self.LABEL_FONT
            ws.cell(row=row + 1, column=2, value=source_filename)

        ws.column_dimensions[get_column_letter(1)].width = 26
        ws.column_dimensions[get_column_letter(2)].width = 60

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()


excel_service = ExcelService()
