"""
Input JSON shape (as produced by the user's OCR tool):
{
  "pages": [
    {
      "page_number": int,
      "text": "...",
      "tables": [ [ [cell, cell, ...], [cell, cell, ...], ... ], ... ],
      "key_value_pairs": [...]
    },
    ...
  ]
}


Extract "List of documents to be scanned and  uploaded" (clause list + any tables present) from a
tender OCR JSON file and write them to Excel.
"""

import json
import re
import sys
from difflib import SequenceMatcher
# from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

FONT_NAME = "Arial"
HEADER_FONT = Font(name=FONT_NAME, bold=True)
BODY_FONT = Font(name=FONT_NAME)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=12)
WRAP = Alignment(wrap_text=True, vertical="top")


TARGET_HEADING = "list documents scanned and uploaded"

def load_pages(json_path: str) -> list:
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return data["pages"]


def _partial_ratio(a: str, b: str) -> float:
    """
    Best-case similarity of `a` against any substring of `b` the same
    length as `a`. Unlike SequenceMatcher.ratio(), this isn't penalized
    when `b` is much longer than `a` (e.g. a heading embedded in a line
    with trailing sentence text).
    """
    if not a or not b:
        return 0.0
    if len(a) > len(b):
        a, b = b, a
    scorer = SequenceMatcher(None, a, b)
    best = 0.0
    for block in scorer.get_matching_blocks():
        start = max(block.b - block.a, 0)
        window = b[start:start + len(a)]
        best = max(best, SequenceMatcher(None, a, window).ratio())
    return best

def find_page(pages: list, heading: str, threshold: float = 0.8) -> tuple:
    """
    Return the first page whose heading matches `heading`.
    Returns (score, page, matched_line).

    Prefers pages[i]["headings"] (paragraphs Document Intelligence already
    classified as role "title"/"sectionHeading") over raw text scanning —
    cleaner and immune to false positives from body text. Falls back to
    fuzzy-matching each page's leading ~20 words only for pages that have
    no "headings" key or an empty one (e.g. OCR JSON without paragraph
    roles, or a doc where Document Intelligence didn't tag anything).
    """

    heading_norm = heading.strip().lower()

    for page in pages:
        for candidate in page.get("headings", []):
            score = _partial_ratio(heading_norm, candidate.strip().lower())
            if score >= threshold:
                return score, page, candidate.strip()


    for page in pages:
        for line in page["text"].splitlines()[:2]: # Look at first 2 lines
            line_norm = line.strip().lower() 
            print("line_norm:", line_norm)
            if not line_norm:
                continue
            # score = SequenceMatcher(None, line_norm, heading_norm).ratio()
            score = _partial_ratio(heading_norm, line_norm)
            if score >= threshold:
                return score, page, line.strip()

    raise ValueError(
        f'Heading "{heading}" not found (>= {threshold:.0%} similarity) '
        f'as a classified heading or in the leading text of any page.'
    )

def extract_clauses(page: dict, matched_heading_line: str) -> list:
    """
    Pull the numbered clause list that follows `matched_heading_line` on the
    page. `matched_heading_line` must be the exact line text as it appears
    in page["text"] (as returned by find_page). Stops as soon as a line
    matches the first header cell of the page's first table, if any table
    exists — so table content never leaks into the clause list.
    """
    lines = [l.strip() for l in page["text"].splitlines() if l.strip()]

    start = next(i for i, l in enumerate(lines) if l == matched_heading_line) + 1

    stop_marker = None
    tables = page.get("tables") or []
    if tables and tables[0] and tables[0][0]:
        stop_marker = tables[0][0][0].strip().lower()

    clause_re = re.compile(r"^(\d+)\.\s*(.*)$")
    clauses = []
    for line in lines[start:]:
        if stop_marker and line.strip().lower() == stop_marker:
            break
        m = clause_re.match(line)
        if m:
            clauses.append([m.group(1), m.group(2)])
        elif clauses:
            clauses[-1][1] = f"{clauses[-1][1]} {line}".strip()

    return clauses


def get_tables(page: dict) -> list:
    """
    Return every table on `page`, whatever their shape or header text.
    Empty list if the page has no tables — never raises.
    """
    tables = page.get("tables") or []
    # drop any malformed/empty tables defensively (e.g. [] or [[]])
    return [t for t in tables if t and t[0]]


def autofit(ws, rows, max_width=70):
    widths = {}
    for row in rows:
        for i, val in enumerate(row, start=1):
            widths[i] = min(max_width, max(widths.get(i, 10), len(str(val)) + 2))
    for i, w in widths.items():
        current = ws.column_dimensions[get_column_letter(i)].width or 0
        ws.column_dimensions[get_column_letter(i)].width = max(current, w)


def write_title(ws, text, row):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = TITLE_FONT
    return row + 1


def write_table(ws, header, rows, start_row=1):
    """Write `header` + `rows` starting at `start_row`. Returns the next free row."""
    for col, val in enumerate(header, start=1):
        cell = ws.cell(row=start_row, column=col, value=val)
        cell.font = HEADER_FONT

    for r, row in enumerate(rows, start=start_row + 1):
        for col, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = BODY_FONT
            cell.alignment = WRAP

    autofit(ws, [header] + rows)
    return start_row + 1 + len(rows)

def extract_data(pages: list, target_heading: str = TARGET_HEADING, threshold: float = 0.8) -> dict:
    """
    Locate `target_heading` in `pages` and return everything that would be
    written to the workbook as a plain, JSON-serializable dict. Inspect
    this (print/json.dumps/assert on it) before touching openpyxl at all.
    """
    score, page, matched_line = find_page(pages, target_heading, threshold=threshold)
    clauses = extract_clauses(page, matched_line)
    tables = get_tables(page)

    return {
        "target_heading": target_heading,
        "matched_heading": matched_line,
        "match_score": round(score, 4),
        "page_number": page["page_number"],
        "clauses": [{"clause_no": no, "text": text} for no, text in clauses],
        "tables": [{"header": table[0], "rows": table[1:]} for table in tables],
    }


def _sanitize_sheet_name(name: str) -> str:
    """Excel sheet names: max 31 chars, no : \\ / ? * [ ]."""
    cleaned = re.sub(r'[:\\/?*\[\]]', "", name)
    return cleaned[:31] or "Sheet1"

def build_workbook(data: dict) -> Workbook:
    """Build (but do not save) a workbook from the dict returned by extract_data()."""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = _sanitize_sheet_name(data["target_heading"])

    ws1.cell(row=1, column=1, value="Target Heading:")
    ws1.cell(row=1, column=2, value=data.get("target_heading", ""))
    ws1.cell(row=2, column=1, value="Matched Heading:")
    ws1.cell(row=2, column=2, value=data.get("matched_heading", ""))
    ws1.cell(row=3, column=1, value="Page Number:")
    ws1.cell(row=3, column=2, value=data.get("page_number", ""))

    clause_rows = [[c["clause_no"], c["text"]] for c in data["clauses"]]
    row = write_table(ws1, ["Clause No.", "Clause Text"], clause_rows, start_row=5)

    for idx, table in enumerate(data["tables"], start=1):
        row = write_title(ws1, f"Table {idx}", row + 2)
        row = write_table(ws1, table["header"], table["rows"], start_row=row)

    return wb    

