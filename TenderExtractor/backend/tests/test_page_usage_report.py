"""Unit tests for the Page Usage Report sheet builder."""
import openpyxl
import pytest

from app import config
from app.pipeline.context import PipelineContext
from app.services import job_store, page_usage_report


@pytest.fixture(autouse=True)
def isolated_db(tmp_path, monkeypatch):
    monkeypatch.setattr(config, "JOBS_DB", tmp_path / "jobs.db")
    monkeypatch.setattr(config, "MAX_PAGES_PER_MONTH", 1000)


def test_build_metrics_reflects_this_months_usage_and_this_jobs_pages():
    # "Tenders scanned"/"Avg pages per tender" count every job recorded
    # this month, not just COMPLETED ones - a failed job still scanned
    # whatever pages it got through before failing.
    job_store.update_job(job_store.new_job_id(), blob_name="a.pdf", status="COMPLETED", ocr_page_count=30)
    job_store.update_job(job_store.new_job_id(), blob_name="b.pdf", status="FAILED", ocr_page_count=5)

    ctx = PipelineContext(blob_name="tender.pdf", job_id="job-x")
    ctx.ocr_page_count = 12

    metrics = dict(page_usage_report.build_metrics(ctx))

    assert metrics["Job ID"] == "job-x"
    assert metrics["Monthly page limit"] == 1000
    assert metrics["Pages scanned this month (all jobs)"] == 35
    assert metrics["Pages remaining this month"] == 965
    assert metrics["Pages scanned by this tender"] == 12
    assert metrics["Tenders scanned this month"] == 2
    assert metrics["Avg. pages scanned per tender this month"] == 17.5


def test_add_usage_sheet_writes_metric_value_table():
    ctx = PipelineContext(blob_name="tender.pdf", job_id="job-x")
    ctx.ocr_page_count = 4

    wb = openpyxl.Workbook()
    page_usage_report.add_usage_sheet(wb, ctx)

    ws = wb[page_usage_report.SHEET_TITLE]
    assert ws.cell(row=1, column=1).value == "Metric"
    assert ws.cell(row=1, column=2).value == "Value"
    labels = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    assert "Job ID" in labels
    assert "Pages scanned by this tender" in labels
