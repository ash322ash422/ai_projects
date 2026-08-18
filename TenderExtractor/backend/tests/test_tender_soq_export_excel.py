"""Unit tests for the standalone Schedule of Quantity Excel writer."""
import openpyxl

from app.services import tender_soq_export_excel as svc


def test_export_writes_table_rows_when_present(tmp_path):
    extraction = {
        "present": True,
        "sections": [{
            "heading": "Schedule of Work",
            "source_pages": [64],
            "tables": [{
                "caption": None,
                "columns": ["S.No.", "Description of Item", "Unit", "Quantity", "Rate", "Amount"],
                "rows": [
                    ["1", "Excavation in ordinary soil", "Cum", "120", "150", "18000"],
                    ["2", "PCC 1:4:8 in foundation", "Cum", "40", "4500", "180000"],
                ],
            }],
            "notes": [],
        }],
    }

    out = svc.export_extraction_to_excel(extraction, tmp_path / "soq.xlsx")

    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == [svc.SHEET_TITLE]
    ws = wb[svc.SHEET_TITLE]
    assert ws.cell(row=1, column=1).value == "Schedule of Work — pages: 64"
    assert ws.cell(row=2, column=1).value == "S.No."
    assert ws.cell(row=3, column=1).value == "1"
    assert ws.cell(row=3, column=2).value == "Excavation in ordinary soil"
    assert ws.cell(row=3, column=4).value == "120"


def test_export_writes_null_placeholder_when_not_present(tmp_path):
    extraction = {"present": False, "sections": []}

    out = svc.export_extraction_to_excel(extraction, tmp_path / "soq.xlsx")

    wb = openpyxl.load_workbook(out)
    ws = wb[svc.SHEET_TITLE]
    assert ws.cell(row=1, column=2).value == "null"


def test_export_writes_multiple_tables_in_one_section(tmp_path):
    extraction = {
        "present": True,
        "sections": [{
            "heading": "Schedule of Quantities",
            "source_pages": [60, 61],
            "tables": [
                {"caption": "Electrical", "columns": ["S.No.", "Quantity"], "rows": [["1", "5"]]},
                {"caption": "Civil", "columns": ["S.No.", "Quantity"], "rows": [["1", "10"]]},
            ],
            "notes": ["Rates exclusive of GST."],
        }],
    }

    out = svc.export_extraction_to_excel(extraction, tmp_path / "soq.xlsx")

    wb = openpyxl.load_workbook(out)
    ws = wb[svc.SHEET_TITLE]
    values = [
        cell.value for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=1) for cell in row
        if cell.value
    ]
    assert "Table: Electrical" in values
    assert "Table: Civil" in values
    assert "Notes" in values
