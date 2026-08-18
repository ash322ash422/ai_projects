# stages.py
"""
Each function here is one pipeline stage. A stage takes the shared
PipelineContext, does one job, and writes its result back onto it.

To add a new stage: write a function with this same signature, then
register it in STAGES in app/pipeline/runner.py. Nothing else needs
to change.
"""
import json
import openpyxl

from app import config
from app.pipeline.context import PipelineContext
from app.pipeline.exceptions import TenderCheckStopped
from app.services import (
    blob_storage,
    document_intelligence,
    page_usage_report,
    tender_detection,
    validation,
    consolidate_excels_files,
    email_service,
)

from app.services.prompt import FIELDS_TO_EXTRACT
from app.utils.logging_config import get_logger
from app.services import (
                        tender_index_extract_service,
                        tender_misc_export_excel,
                        tender_misc_extraction_service,
                        tender_nit_export_excel,
                        tender_nit_extract_service,
                        tender_soq_export_excel,
                        tender_soq_extract_service,
)

logger = get_logger(__name__)


def _accumulate_tokens(ctx: PipelineContext, count: int) -> None:
    ctx.token_count += count


def _save_json(path, data) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def _load_json(path) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def _get_pages(ctx: PipelineContext) -> list:
    """OCR pages for `ctx` - requires stage_ocr to have already populated ctx.document_data."""
    if ctx.document_data is None:
        raise ValueError(f"No OCR data available for {ctx.blob_name} - stage_ocr must run first.")
    return ctx.document_data["pages"]


def _get_preview_pages(ctx: PipelineContext) -> list:
    """Preview OCR pages for `ctx` - requires stage_ocr_preview to have already populated
    ctx.preview_document_data."""
    if ctx.preview_document_data is None:
        raise ValueError(f"No preview OCR data available for {ctx.blob_name} - stage_ocr_preview must run first.")
    return ctx.preview_document_data["pages"]


# --------------------------------------------------------------------------
# Stage 1: get the PDF bytes onto local disk / into memory
# --------------------------------------------------------------------------
def stage_ingest(ctx: PipelineContext) -> None:
    if ctx.local_pdf_path.exists():
        # Already placed in this job's own folder - the API writes
        # uploads directly here (see app/api/main.py: create_tender).
        ctx.pdf_bytes = ctx.local_pdf_path.read_bytes()
    else:
        # CLI flow (main.py): PDFs get dropped flat into data_uploads/
        # before any job_id exists. Read from there and copy into this
        # job's own folder, so every run - CLI or API - ends up with
        # an isolated, immutable copy of exactly what it processed,
        # and concurrent jobs for a same-named file can't collide.
        legacy_path = config.DATA_UPLOAD_DIR / ctx.blob_name
        if not legacy_path.exists():
            raise FileNotFoundError(f"PDF not found: {legacy_path}")
        ctx.pdf_bytes = legacy_path.read_bytes()
        ctx.local_pdf_path.parent.mkdir(parents=True, exist_ok=True)
        ctx.local_pdf_path.write_bytes(ctx.pdf_bytes)


# --------------------------------------------------------------------------
# Stage 2a: OCR just the first config.INDEX_CHECK_PAGES pages. Cheap enough
# to run before committing to a full-document OCR pass - lets
# stage_check_index_and_tender rule out a non-tender upload (and the LLM
# call it would otherwise trigger) without paying for the whole document.
# --------------------------------------------------------------------------
def stage_ocr_preview(ctx: PipelineContext) -> None:
    document_intelligence.validate_pdf_is_digital(ctx.pdf_bytes)

    result = document_intelligence.analyze_document_from_bytes(
        document_bytes=ctx.pdf_bytes,
        endpoint=config.AZURE_DOCINTEL_ENDPOINT,
        key=config.AZURE_DOCINTEL_KEY,
        model_name=config.AZURE_DOCINTEL_MODEL,
        pages=f"1-{config.INDEX_CHECK_PAGES}",
    )

    ctx.preview_document_data = document_intelligence.extract_document_data(result)
    ctx.ocr_page_count = len(ctx.preview_document_data.get("pages", []))
    logger.info("[%s] Preview OCR processed %d page(s)", ctx.blob_name, ctx.ocr_page_count)

    _save_json(ctx.job_audit_path("ocr_preview.json"), ctx.preview_document_data)


# --------------------------------------------------------------------------
# Stage: extract "index" (table of contents) data, and check whether this
# document is actually a tender at all (keyword match on the same preview
# pages). Raises TenderCheckStopped - caught specially by run_pipeline - if
# it isn't, so the (expensive) full OCR pass below never runs for it.
# --------------------------------------------------------------------------
def stage_check_index_and_tender(ctx: PipelineContext) -> None:
    pages = _get_preview_pages(ctx)

    ctx.is_tender = tender_detection.looks_like_tender(pages)
    if not ctx.is_tender:
        raise TenderCheckStopped(
            f"'{ctx.blob_name}' does not appear to be a tender - none of "
            f"{tender_detection.TENDER_KEYWORDS} found in the first {config.INDEX_CHECK_PAGES} page(s)."
        )

    try:
        ctx.extracted_index_data = tender_index_extract_service.extract_data(
            pages, token_callback=lambda n: _accumulate_tokens(ctx, n)
        )
        _save_json(ctx.job_audit_path("index.json"), ctx.extracted_index_data)

    except ValueError:
        logger.warning(
            "[%s] '%s' section not found, skipping.",
            ctx.blob_name, tender_index_extract_service.TARGET_HEADING,
        )


# --------------------------------------------------------------------------
# Stage 2b: full-document OCR / layout extraction, for confirmed tenders
# only. Result is stored in ctx.document_data and written to this job's
# audit copy under logs/{job_id}/.
# --------------------------------------------------------------------------
def stage_ocr(ctx: PipelineContext) -> None:
    result = document_intelligence.analyze_document_from_bytes(
        document_bytes=ctx.pdf_bytes,
        endpoint=config.AZURE_DOCINTEL_ENDPOINT,
        key=config.AZURE_DOCINTEL_KEY,
        model_name=config.AZURE_DOCINTEL_MODEL,
    )

    ctx.document_data = document_intelligence.extract_document_data(result)
    # += : stage_ocr_preview already counted the first config.INDEX_CHECK_PAGES
    # pages against this job - this full pass re-scans them too (accepted
    # simplicity over merging the two OCR results), so both counts add up.
    ctx.ocr_page_count += len(ctx.document_data.get("pages", []))
    logger.info("[%s] OCR processed %d page(s)", ctx.blob_name, len(ctx.document_data.get("pages", [])))

    _save_json(ctx.job_audit_path("ocr.json"), ctx.document_data)



# --------------------------------------------------------------------------
# Stage 3: Field extraction via LLM: Loads the OCRed text data and then 
# sends it to LLM for extraction of NIT data
# --------------------------------------------------------------------------
def stage_extract_nit_data(ctx: PipelineContext) -> None:
    ctx.extracted_nit_data = tender_nit_extract_service.extract_document(
        document_data=ctx.document_data,
        fields=FIELDS_TO_EXTRACT,
        token_callback=lambda n: _accumulate_tokens(ctx, n),
    )
    _save_json(ctx.job_audit_path("extracted_nit.json"), ctx.extracted_nit_data)


# --------------------------------------------------------------------------
# Stage 4: validate / normalize the NIT data
# --------------------------------------------------------------------------
def stage_validate_nit(ctx: PipelineContext) -> None:
    ctx.validated_nit_data = validation.validate_nit_extracted_json(ctx.extracted_nit_data)
    _save_json(ctx.job_audit_path("validated_nit.json"), ctx.validated_nit_data)


# --------------------------------------------------------------------------
# Stage 5: Export to Excel
# --------------------------------------------------------------------------
def stage_export_nit_data(ctx: PipelineContext) -> None:
    audit_df = tender_nit_export_excel.validation_json_to_dataframe(ctx.validated_nit_data)
    tender_nit_export_excel.save_dataframe_to_excel(audit_df, ctx.audit_nit_excel_path)

    clean_df = tender_nit_export_excel.validation_json_to_clean_dataframe(ctx.validated_nit_data)
    tender_nit_export_excel.save_dataframe_to_excel(clean_df, ctx.clean_nit_excel_path)


# --------------------------------------------------------------------------
# Stage : extract "terms and conditions", "List of document to be  scanned and upload", "Acceptable make"
# --------------------------------------------------------------------------
def stage_extract_tender_misc_data(ctx: PipelineContext) -> None:
    pages = _get_pages(ctx)

    try:
        ctx.extracted_misc_data = tender_misc_extraction_service.extract_data(
            pages,
            token_callback=lambda n: _accumulate_tokens(ctx, n),
            page_chunk_size=config.MISC_PAGES_PER_CHUNK,
        )
        _save_json(ctx.job_audit_path("misc.json"), ctx.extracted_misc_data)

    except ValueError:
        logger.warning("[%s] Miscellaneous data not found", ctx.blob_name)


# --------------------------------------------------------------------------
# Stage: Export the miscellaneous data to excel
# --------------------------------------------------------------------------
def stage_export_tender_misc_data(ctx: PipelineContext) -> None:

    data = ctx.extracted_misc_data

    # 1. Defensive Guard Clause
    if not data:
        logger.warning("[%s] No extracted misc. data found in context. Skipping export.", 
                       ctx.blob_name
        )
        return

    # 2. Safe Generation and Save
    try:
        filename = tender_misc_export_excel.export_extraction_to_excel(data, ctx.exported_misc_excel_path)
        logger.info("[%s] Wrote misc data to -> %s", 
                    ctx.blob_name, ctx.exported_misc_excel_path)
    except (ValueError, KeyError, TypeError) as e:
            logger.error(
                "[%s] Failed to build or save misc workbook. Error: %s",
                ctx.blob_name, str(e)
            )



# --------------------------------------------------------------------------
# Stage: extract "Schedule of Quantity" from just the last config.SOQ_LAST_PAGES
# pages - kept separate from stage_extract_tender_misc_data (see
# tender_soq_extract_service.py for why: it lives in a predictable spot near
# the end of the document, so a narrow, focused call is cheaper and less
# prone to misclassifying unrelated content than scanning the whole thing).
# --------------------------------------------------------------------------
def stage_extract_schedule_of_quantity(ctx: PipelineContext) -> None:
    pages = _get_pages(ctx)
    last_pages = pages[-config.SOQ_LAST_PAGES:]

    try:
        ctx.extracted_soq_data = tender_soq_extract_service.extract_data(
            last_pages, token_callback=lambda n: _accumulate_tokens(ctx, n)
        )
        _save_json(ctx.job_audit_path("soq.json"), ctx.extracted_soq_data)

    except ValueError:
        logger.warning("[%s] Schedule of Quantity not found", ctx.blob_name)


# --------------------------------------------------------------------------
# Stage: Export the Schedule of Quantity data to excel
# --------------------------------------------------------------------------
def stage_export_schedule_of_quantity(ctx: PipelineContext) -> None:
    data = ctx.extracted_soq_data

    if not data:
        logger.warning("[%s] No extracted Schedule of Quantity data found in context. Skipping export.",
                       ctx.blob_name
        )
        return

    try:
        tender_soq_export_excel.export_extraction_to_excel(data, ctx.exported_soq_excel_path)
        logger.info("[%s] Wrote schedule of quantity data to -> %s",
                    ctx.blob_name, ctx.exported_soq_excel_path)
    except (ValueError, KeyError, TypeError) as e:
        logger.error(
            "[%s] Failed to build or save schedule of quantity workbook. Error: %s",
            ctx.blob_name, str(e)
        )


# --------------------------------------------------------------------------
# Stage 8: consolidate all excels into one workbook,
# --------------------------------------------------------------------------
def stage_consolidate_all_excels(ctx: PipelineContext) -> None:
    # (file_path, sheet_name_override): override renames a single-sheet file's
    # sheet; None means "copy every sheet from this workbook, keep its own name".
    input_files = [
        (ctx.clean_nit_excel_path, "NIT Data"),
        (ctx.exported_misc_excel_path, None),
        (ctx.exported_soq_excel_path, None),
    ]
    
    existing_files = [(p, override) for p, override in input_files if p.exists()]
    if not existing_files:
        logger.warning("[%s] No individual excel outputs found to consolidate. Skipping.",
                       ctx.blob_name
        )
        return
    if len(existing_files) < len(input_files):
        existing_paths = {p for p, _ in existing_files}
        missing = [p.name for p, _ in input_files if p not in existing_paths]
        logger.warning("[%s] Missing excel outputs, consolidating available ones only: %s",
                       ctx.blob_name, missing
        )

    try:
        consolidated_wb = openpyxl.Workbook()
        consolidated_wb.remove(consolidated_wb.active)  # drop the default blank sheet

        for path, override in existing_files:
            src_wb = openpyxl.load_workbook(path)
            # override => only that file's first sheet, renamed; None => every sheet, original names
            src_sheet_names = [src_wb.sheetnames[0]] if override else src_wb.sheetnames
            for src_sheet_name in src_sheet_names:
                src_ws = src_wb[src_sheet_name]
                title = (override or src_sheet_name)[:31]  # Excel sheet names capped at 31 chars
                dst_ws = consolidated_wb.create_sheet(title)
                consolidate_excels_files._copy_worksheet(src_ws, dst_ws)
            src_wb.close()

        ctx.consolidated_excel_path.parent.mkdir(parents=True, exist_ok=True)
        consolidated_wb.save(ctx.consolidated_excel_path)
        logger.info("[%s] Wrote consolidated excel -> %s",
                    ctx.blob_name,
                    ctx.consolidated_excel_path
        )
    except (ValueError, KeyError, TypeError, OSError) as e:
        logger.error(
            "[%s] Failed to build or save consolidated workbook. Error: %s",
            ctx.blob_name, str(e),
        )


# --------------------------------------------------------------------------
# Stage: append a "Page Usage Report" sheet (monthly OCR page quota vs.
# actual usage) to the consolidated workbook before it's published.
# --------------------------------------------------------------------------
def stage_add_page_usage_report(ctx: PipelineContext) -> None:
    if not ctx.consolidated_excel_path.exists():
        logger.warning("[%s] No consolidated excel found, skipping page usage report.",
                       ctx.blob_name
        )
        return

    try:
        wb = openpyxl.load_workbook(ctx.consolidated_excel_path)
        page_usage_report.add_usage_sheet(wb, ctx)
        wb.save(ctx.consolidated_excel_path)
        logger.info("[%s] Added page usage report sheet -> %s",
                    ctx.blob_name, ctx.consolidated_excel_path
        )
    except (ValueError, KeyError, TypeError, OSError) as e:
        logger.error(
            "[%s] Failed to add page usage report sheet. Error: %s",
            ctx.blob_name, str(e),
        )


# --------------------------------------------------------------------------
# Stage 10: publish the output (Blob Storage if configured, else local disk)
# --------------------------------------------------------------------------
def stage_publish(ctx: PipelineContext) -> None:
    if not config.AZURE_STORAGE_CONNECTION_STRING:
        ctx.download_url = str(ctx.consolidated_excel_path)
        logger.info("[%s] Blob Storage not configured - output kept locally at %s", 
                    ctx.blob_name, 
                    ctx.download_url
        )
        return

    client = blob_storage.get_blob_service_client(config.AZURE_STORAGE_CONNECTION_STRING)
    processed_container = blob_storage.get_or_create_container(client,
                                                               config.BLOB_CONTAINER_PROCESSED
    )

    # Job-scoped blob name so two jobs producing a same-named output (e.g.
    # two runs of "tender.xlsx") never overwrite each other in the
    # processed container or hand out a download URL for the wrong job's
    # file - the same collision this whole change fixes on local disk.
    processed_blob_name = f"{ctx.job_id}/{ctx.consolidated_excel_path.name}"

    blob_storage.upload_file(processed_container, ctx.consolidated_excel_path, processed_blob_name)

    ctx.download_url = blob_storage.generate_download_url(
        connection_string=config.AZURE_STORAGE_CONNECTION_STRING,
        container_name=config.BLOB_CONTAINER_PROCESSED,
        blob_name=processed_blob_name,
        expiry_hours=config.URL_EXPIRY_HOURS,
    )

# --------------------------------------------------------------------------
# Stage 11: notify (optional, non-fatal - a failed email shouldn't fail a
# tender that was otherwise processed successfully)
# --------------------------------------------------------------------------
def stage_notify(ctx: PipelineContext) -> None:
    if not config.NOTIFY_ON_COMPLETE:
        return

    if len(config.NOTIFY_RECIPIENTS) == 0:
        logger.warning("[%s] NOTIFY_ON_COMPLETE is set but NOTIFY_RECIPIENTS is empty, skipping.",
                       ctx.blob_name
        )
        return

    for recipient in config.NOTIFY_RECIPIENTS:
        try:
            email_service.send_notification(recipient, 
                                            ctx.download_url, 
                                            ctx.blob_name
            )
        except Exception:
            logger.exception("[%s] Notification email failed, tender output is still valid.",
                            ctx.blob_name
            )
