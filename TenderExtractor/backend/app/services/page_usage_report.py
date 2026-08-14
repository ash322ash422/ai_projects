# page_usage_report.py
"""
Builds the "Page Usage Report" sheet appended to every tender's
consolidated Excel deliverable: the shared monthly OCR page quota
(config.MAX_PAGES_PER_MONTH) against actual usage, so a user checking one
tender's output can also see where the shared budget stands.
"""
from __future__ import annotations

from datetime import datetime
from typing import TYPE_CHECKING

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app import config
from app.services import job_store

if TYPE_CHECKING:
    from app.pipeline.context import PipelineContext

SHEET_TITLE = "Page Usage Report"

_FONT = "Arial"
_HEADER_FONT = Font(name=_FONT, bold=True, size=11)
_LABEL_FONT = Font(name=_FONT, bold=True, size=11)
_VALUE_FONT = Font(name=_FONT, size=11)
_HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")


def build_metrics(ctx: "PipelineContext") -> list[tuple[str, object]]:
    now = datetime.now(config.IST)
    limit = config.MAX_PAGES_PER_MONTH

    jobs = job_store.jobs_this_month(now)
    used = sum(job["ocr_page_count"] for job in jobs)
    tender_count = len(jobs)

    return [
        ("Job ID", ctx.job_id),
        ("Report generated (IST)", now.strftime("%Y-%m-%d %H:%M:%S")),
        ("Reporting month", now.strftime("%Y-%m")),
        ("Monthly page limit", limit),
        ("Pages scanned this month (all jobs)", used),
        ("Pages remaining this month", max(limit - used, 0)),
        ("% of monthly limit used", round(100 * used / limit, 1) if limit else None),
        ("Pages scanned by this tender", ctx.ocr_page_count),
        ("Tenders scanned this month", tender_count),
        (
            "Avg. pages scanned per tender this month",
            round(used / tender_count, 1) if tender_count else None,
        ),
    ]


def add_usage_sheet(wb: openpyxl.Workbook, ctx: "PipelineContext") -> Worksheet:
    ws = wb.create_sheet(SHEET_TITLE[:31])  # Excel sheet name limit

    ws.cell(row=1, column=1, value="Metric").font = _HEADER_FONT
    ws.cell(row=1, column=1).fill = _HEADER_FILL
    ws.cell(row=1, column=2, value="Value").font = _HEADER_FONT
    ws.cell(row=1, column=2).fill = _HEADER_FILL

    for row, (label, value) in enumerate(build_metrics(ctx), start=2):
        ws.cell(row=row, column=1, value=label).font = _LABEL_FONT
        ws.cell(row=row, column=1).alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row=row, column=2, value=value).font = _VALUE_FONT

    ws.column_dimensions[get_column_letter(1)].width = 45
    ws.column_dimensions[get_column_letter(2)].width = 20
    ws.freeze_panes = "A2"
    return ws
