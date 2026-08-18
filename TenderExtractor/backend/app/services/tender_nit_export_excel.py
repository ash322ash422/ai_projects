"""
Turns validated field data into the two Excel deliverables:
  - an audit workbook (original + normalized + valid flag per field)
  - a clean workbook (one tidy value per field, ready to hand off)
"""

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pathlib import Path

_FONT = "Arial"
_HEADER_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # light green, matches the table-header color used elsewhere
_HEADER_FONT = Font(name=_FONT, bold=True, size=11)
_BODY_FONT = Font(name=_FONT, size=11)
_WRAP = Alignment(wrap_text=True, vertical="top")
_WRAP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")

_MIN_COL_WIDTH = 12
_MAX_COL_WIDTH = 60


def validation_json_to_dataframe(data: dict) -> pd.DataFrame:
    row = {}
    for field, info in data.items():
        row[f"{field}_original"] = info.get("original")
        row[f"{field}_normalized"] = info.get("normalized")
        row[f"{field}_valid"] = info.get("valid")
    return pd.DataFrame([row])


def validation_json_to_clean_dataframe(data: dict) -> pd.DataFrame:
    row = {}
    for field, info in data.items():
        # row[field] = info.get("normalized") if info.get("valid", False) else info.get("original") # Use this line if you want to keep the original value when the normalized value is invalid
        row[field] = info.get("original") # Use this line if you want to keep the original value regardless of validity
        
    return pd.DataFrame([row])


def save_dataframe_to_excel(df: pd.DataFrame, excel_file: Path) -> None:
    excel_file.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = (excel_file.stem or "Sheet1")[:31]  # Excel sheet names capped at 31 chars

    for c, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=c, value=str(col_name))
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _WRAP_CENTER

    for r, row in enumerate(df.itertuples(index=False), start=2):
        for c, value in enumerate(row, start=1):
            if pd.isna(value):
                value = None
            cell = ws.cell(row=r, column=c, value=value)
            cell.font = _BODY_FONT
            cell.alignment = _WRAP

    # Size each column to its content (header included) instead of Excel's default
    # auto-width, which is what was causing headers to visually spill into the
    # next column; capped so a single long cell can't blow out the whole sheet.
    for c, col_name in enumerate(df.columns, start=1):
        # Compute lengths value-by-value with plain str() rather than
        # Series.astype(str) -- on some pandas versions/dtypes (nullable
        # Float64/Int64, certain object columns), astype(str) can leave NaN/NA
        # entries un-stringified, which then blows up len(). Skipping missing
        # values here also means a mostly-empty column doesn't get stretched
        # wide for no reason.
        lengths = [len(str(col_name))]
        for value in df.iloc[:, c - 1]:
            if pd.isna(value):
                continue
            lengths.append(len(str(value)))
        longest = max(lengths)
        ws.column_dimensions[get_column_letter(c)].width = min(
            max(longest + 2, _MIN_COL_WIDTH), _MAX_COL_WIDTH
        )

    ws.freeze_panes = "A2"
    wb.save(excel_file)