# tender_soq_export_excel.py
"""
Renders Schedule of Quantity extraction (see tender_soq_extract_service.py)
to a standalone one-sheet Excel workbook. Later merged into the
consolidated deliverable by stage_consolidate_all_excels - the same
extract-service/export-service split used for NIT data
(tender_nit_extract_service.py / tender_nit_export_excel.py).

Self-contained (its own style constants, not imported from
tender_misc_export_excel.py) so this stays independently readable and
doesn't reach into another module's private names for what's ultimately a
handful of duplicated lines.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.utils.logging_config import get_logger

logger = get_logger(__name__)

SHEET_TITLE = "Schedule of Quantity"

_FONT = "Arial"
_SUBHEADER_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
_TABLE_HEADER_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
_HEADER_FONT = Font(name=_FONT, bold=True, size=11)
_SUBHEADER_FONT = Font(name=_FONT, bold=True, italic=True, size=11)
_BODY_FONT = Font(name=_FONT, size=11)
_ITALIC_FONT = Font(name=_FONT, italic=True, size=10)
_WRAP = Alignment(wrap_text=True, vertical="top")
_WRAP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")

_MIN_COL_WIDTH = 10
_MAX_COL_WIDTH = 60
_NUMBER_COL_WIDTH = 10


def export_extraction_to_excel(extraction: dict, output_path) -> str:
    """
    Convert the dict returned by extract_data() (matching the
    ScheduleOfQuantityExtraction schema: {"present": bool, "sections":
    [...]}) into a single-sheet Excel workbook.

    A "present": False extraction produces a sheet stating the value is
    null, mirroring extract_data()'s "not found" result rather than
    leaving the sheet blank or raising an error.

    Args:
        extraction: dict as returned by tender_soq_extract_service.extract_data(...).
        output_path: path to write the .xlsx file to.

    Returns:
        The output_path that was written, for convenience chaining.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # replace the default blank sheet with our own

    ws = wb.create_sheet(SHEET_TITLE[:31])  # Excel sheet name limit
    _write_sheet(ws, extraction)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info("Wrote schedule of quantity workbook to %s", output_path)
    return str(output_path)


def _write_sheet(ws: Worksheet, extraction: dict) -> None:
    max_cols_used = 2
    r = 1

    if not extraction.get("present"):
        ws.cell(row=r, column=1, value="Value:").font = _HEADER_FONT
        ws.cell(row=r, column=2, value="null").font = _ITALIC_FONT
        r += 2
        ws.cell(row=r, column=1, value="Note").font = Font(name=_FONT, bold=True, italic=True, size=10)
        ws.cell(
            row=r, column=2,
            value="'schedule_of_quantity' was not found anywhere in the provided document text.",
        ).font = _ITALIC_FONT
        ws.cell(row=r, column=2).alignment = _WRAP
        _set_column_widths(ws, max_cols_used)
        return

    sections = extraction.get("sections") or []
    if not sections:
        ws.cell(row=r, column=1, value="Value:").font = _HEADER_FONT
        ws.cell(row=r, column=2, value="present, but no sections were returned").font = _ITALIC_FONT
        _set_column_widths(ws, max_cols_used)
        return

    for section in sections:
        r, section_max_cols = _write_section(ws, r, section)
        max_cols_used = max(max_cols_used, section_max_cols)
        r += 1  # blank row between sections

    ws.freeze_panes = "A2"
    _set_column_widths(ws, max_cols_used)


def _write_section(ws: Worksheet, start_row: int, section: dict) -> tuple[int, int]:
    r = start_row
    max_cols = 2

    heading = section.get("heading") or "(no heading)"
    pages = section.get("source_pages") or []
    pages_str = f" — pages: {', '.join(str(p) for p in pages)}" if pages else ""
    _write_merged_subheader(ws, r, 6, f"{heading}{pages_str}")
    r += 1

    tables = section.get("tables") or []
    for table in tables:
        table_cols = table.get("columns") or []
        caption = table.get("caption")
        if caption:
            ws.cell(row=r, column=1, value=f"Table: {caption}").font = _SUBHEADER_FONT
            r += 1
        if table_cols:
            for c, col_name in enumerate(table_cols, start=1):
                cell = ws.cell(row=r, column=c, value=col_name)
                cell.font = _HEADER_FONT
                cell.fill = _TABLE_HEADER_FILL
                cell.alignment = _WRAP_CENTER
            max_cols = max(max_cols, len(table_cols))
            r += 1
        for row_values in table.get("rows") or []:
            for c, value in enumerate(row_values, start=1):
                cell = ws.cell(row=r, column=c, value=value)
                cell.font = _BODY_FONT
                cell.alignment = _WRAP_CENTER if c == 1 else _WRAP
            max_cols = max(max_cols, len(row_values))
            r += 1
        r += 1  # blank row between tables within the same section

    notes = section.get("notes") or []
    if notes:
        ws.cell(row=r, column=1, value="Notes").font = Font(name=_FONT, bold=True, italic=True, size=10)
        r += 1
        for note in notes:
            ws.cell(row=r, column=1, value="*").font = _ITALIC_FONT
            ws.cell(row=r, column=1).alignment = _WRAP_CENTER
            ws.cell(row=r, column=2, value=note).font = _ITALIC_FONT
            ws.cell(row=r, column=2).alignment = _WRAP
            r += 1

    return r, max_cols


def _write_merged_subheader(ws: Worksheet, row: int, span_cols: int, text: str) -> None:
    ws.cell(row=row, column=1, value=text)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
    cell = ws.cell(row=row, column=1)
    cell.font = _SUBHEADER_FONT
    cell.fill = _SUBHEADER_FILL
    cell.alignment = Alignment(vertical="center")


def _set_column_widths(ws: Worksheet, max_cols_used: int) -> None:
    ws.column_dimensions["A"].width = _NUMBER_COL_WIDTH
    for c in range(2, max(max_cols_used, 2) + 1):
        ws.column_dimensions[get_column_letter(c)].width = _MAX_COL_WIDTH if c == 2 else _MIN_COL_WIDTH + 20
