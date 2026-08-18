from __future__ import annotations

from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.utils.logging_config import get_logger

logger = get_logger(__name__)



# extraction dict keys -> sheet titles, in the order sheets should appear.
# Schedule of Quantity is NOT here - it's extracted and exported separately
# (see tender_soq_extract_service.py / tender_soq_export_excel.py) and
# merged into the final workbook by stage_consolidate_all_excels instead.
FIELD_TO_SHEET_TITLE = {
    "terms_and_conditions": "Terms and Conditions",
    "acceptable_make": "Acceptable Make",
    "documents_to_be_scanned_and_uploaded": "Documents to Upload",
}

_FONT = "Arial"
_HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
_SUBHEADER_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
_TABLE_HEADER_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
_HEADER_FONT = Font(name=_FONT, bold=True, size=11)
_SUBHEADER_FONT = Font(name=_FONT, bold=True, italic=True, size=11)
_BODY_FONT = Font(name=_FONT, size=11)
_ITALIC_FONT = Font(name=_FONT, italic=True, size=10)
_WRAP = Alignment(wrap_text=True, vertical="top")
_WRAP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")

_MIN_COL_WIDTH = 10
_MAX_COL_WIDTH = 60
_NUMBER_COL_WIDTH = 10


def export_extraction_to_excel(
    extraction: dict,
    output_path: str,
    sheet_titles: Optional[dict] = None,
) -> str:
    """
    Convert the dict returned by extract_data() (matching the TenderExtraction
    schema: {field: {"present": bool, "sections": [...]}}) into an Excel
    workbook with one sheet per field - Terms and Conditions, Acceptable
    Make, Documents to be Scanned and Uploaded.

    Each section within a field is rendered as:
      - a sub-header row with the section's heading (or "(no heading)") and
        its source page numbers, if any
      - its numbered/lettered items as a two-column (No. / Text) block
      - any tables as their own header+rows block, using the table's own
        columns (tables with different column counts are handled
        independently, so a Rate table and a Make table in the same section
        each get their own correctly-shaped block)
      - any notes as an italic block at the end of the section

    A field with "present": False produces a sheet stating the value is
    null, mirroring extract_data()'s "not found" result rather than leaving
    the sheet blank or raising an error.

    Args:
        extraction: dict as returned by extract_data(...) (or
            TenderExtraction(...).model_dump(mode="json")).
        output_path: path to write the .xlsx file to.
        sheet_titles: optional override of {field_key: sheet_title}; falls
            back to FIELD_TO_SHEET_TITLE for any key not supplied.

    Returns:
        The output_path that was written, for convenience chaining.
    """
    titles = {**FIELD_TO_SHEET_TITLE, **(sheet_titles or {})}

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # replace the default blank sheet with our own, in order

    for field_key, sheet_title in titles.items():
        if field_key not in extraction:
            logger.warning("extraction dict is missing expected key %r; skipping sheet.", field_key)
            continue
        ws = wb.create_sheet(sheet_title[:31])  # Excel sheet name limit
        _write_field_sheet(ws, extraction[field_key], field_key)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info("Wrote tender extraction workbook to %s", output_path)
    return output_path


def _write_field_sheet(ws: Worksheet, field: dict, field_key: str) -> None:
    max_cols_used = 2
    r = 1

    if not field.get("present"):
        ws.cell(row=r, column=1, value="Value:").font = _HEADER_FONT
        ws.cell(row=r, column=2, value="null").font = _ITALIC_FONT
        r += 2
        ws.cell(row=r, column=1, value="Note").font = Font(name=_FONT, bold=True, italic=True, size=10)
        ws.cell(
            row=r, column=2,
            value=f"'{field_key}' was not found anywhere in the provided document text.",
        ).font = _ITALIC_FONT
        ws.cell(row=r, column=2).alignment = _WRAP
        _set_column_widths(ws, max_cols_used)
        return

    sections = field.get("sections") or []
    if not sections:
        ws.cell(row=r, column=1, value="Value:").font = _HEADER_FONT
        ws.cell(row=r, column=2, value="present, but no sections were returned").font = _ITALIC_FONT
        _set_column_widths(ws, max_cols_used)
        return

    for section in sections:
        r, section_max_cols = _write_section(ws, r, section)
        max_cols_used = max(max_cols_used, section_max_cols)
        r += 1  # blank row between sections

    ws.freeze_panes = "A2"
    _set_column_widths(ws, max_cols_used)


def _write_section(ws: Worksheet, start_row: int, section: dict) -> tuple[int, int]:
    r = start_row
    max_cols = 2

    heading = section.get("heading") or "(no heading)"
    pages = section.get("source_pages") or []
    pages_str = f" — pages: {', '.join(str(p) for p in pages)}" if pages else ""
    _write_merged_subheader(ws, r, 6, f"{heading}{pages_str}")
    r += 1

    items = section.get("items") or []
    if items:
        ws.cell(row=r, column=1, value="No.").font = _HEADER_FONT
        ws.cell(row=r, column=1).fill = _HEADER_FILL
        ws.cell(row=r, column=1).alignment = _WRAP_CENTER
        ws.cell(row=r, column=2, value="Text").font = _HEADER_FONT
        ws.cell(row=r, column=2).fill = _HEADER_FILL
        ws.cell(row=r, column=2).alignment = _WRAP_CENTER
        r += 1
        for item in items:
            ws.cell(row=r, column=1, value=item.get("number")).font = _BODY_FONT
            ws.cell(row=r, column=1).alignment = _WRAP_CENTER
            ws.cell(row=r, column=2, value=item.get("text", "")).font = _BODY_FONT
            ws.cell(row=r, column=2).alignment = _WRAP
            r += 1

    tables = section.get("tables") or []
    for table in tables:
        r += 1
        table_cols = table.get("columns") or []
        caption = table.get("caption")
        if caption:
            ws.cell(row=r, column=1, value=f"Table: {caption}").font = _SUBHEADER_FONT
            r += 1
        if table_cols:
            for c, col_name in enumerate(table_cols, start=1):
                cell = ws.cell(row=r, column=c, value=col_name)
                cell.font = _HEADER_FONT
                cell.fill = _TABLE_HEADER_FILL
                cell.alignment = _WRAP_CENTER
            max_cols = max(max_cols, len(table_cols))
            r += 1
        for row_values in table.get("rows") or []:
            for c, value in enumerate(row_values, start=1):
                cell = ws.cell(row=r, column=c, value=value)
                cell.font = _BODY_FONT
                cell.alignment = _WRAP_CENTER if c == 1 else _WRAP
            max_cols = max(max_cols, len(row_values))
            r += 1

    notes = section.get("notes") or []
    if notes:
        r += 1
        ws.cell(row=r, column=1, value="Notes").font = Font(name=_FONT, bold=True, italic=True, size=10)
        r += 1
        for note in notes:
            ws.cell(row=r, column=1, value="*").font = _ITALIC_FONT
            ws.cell(row=r, column=1).alignment = _WRAP_CENTER
            ws.cell(row=r, column=2, value=note).font = _ITALIC_FONT
            ws.cell(row=r, column=2).alignment = _WRAP
            r += 1

    return r, max_cols


def _write_merged_subheader(ws: Worksheet, row: int, span_cols: int, text: str) -> None:
    ws.cell(row=row, column=1, value=text)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
    cell = ws.cell(row=row, column=1)
    cell.font = _SUBHEADER_FONT
    cell.fill = _SUBHEADER_FILL
    cell.alignment = Alignment(vertical="center")


def _set_column_widths(ws: Worksheet, max_cols_used: int) -> None:
    ws.column_dimensions["A"].width = _NUMBER_COL_WIDTH
    for c in range(2, max(max_cols_used, 2) + 1):
        ws.column_dimensions[get_column_letter(c)].width = _MAX_COL_WIDTH if c == 2 else _MIN_COL_WIDTH + 20
